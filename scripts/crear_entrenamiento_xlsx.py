"""Genera entrenamiento.xlsx a partir de entrenamiento.jsonl.

Aplanado COMPLETO del schema v1.1.0: una fila por evento de protesta,
con TODOS los campos del json en columnas planas (sin anidacion).
Las notas sin eventos producen una fila unica con las columnas de evento
vacias.

Arrays de objetos (sujetos, motivos, demandas, contra_quien, lugares,
formatos_complementarios, personas, citas, control_extraccion.*) se aplanan
en UNA celda con todos sus pares clave=valor concatenados, separados por
'; ' entre elementos y ' | ' entre pares dentro de un mismo elemento.
Si una nota no tiene el array, la celda queda vacia.

Las columnas conservan los nombres pedidos originalmente (nota_id,
estado_validacion_humana, tags_topicos, modificado_por, es_evento_protesta,
tipo_temporal, accion_valor_textual, accion_categoria, texto_original)
mas todas las demas.
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "entrenamiento.jsonl"
OUT = ROOT / "entrenamiento.xlsx"


def flatten_value(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (int, float)):
        return str(v)
    if isinstance(v, str):
        return v
    if isinstance(v, dict):
        parts = [f"{k}={flatten_value(vv)}" for k, vv in v.items()]
        return " | ".join(parts)
    if isinstance(v, list):
        if not v:
            return ""
        return " ; ".join(flatten_value(x) for x in v)
    return str(v)


def get_path(obj: Any, path: str) -> Any:
    cur = obj
    for part in path.split("."):
        if not isinstance(cur, dict):
            return None
        if part not in cur:
            return None
        cur = cur[part]
    return cur


def array_field(parent: Any, key: str) -> str:
    if not isinstance(parent, dict):
        return ""
    arr = parent.get(key)
    if not isinstance(arr, list) or not arr:
        return ""
    return " ; ".join(flatten_value(item) for item in arr)


def compute_estado(validacion: dict | None) -> str:
    if validacion is None:
        return "Validado sin edicion"
    if validacion.get("modificada"):
        return "Validado con edicion"
    return "Validado sin edicion"


def compute_modificado_por(validacion: dict | None) -> str:
    if validacion is None:
        return ""
    if validacion.get("modificada"):
        return str(validacion.get("usuario_validador") or "Nico")
    return ""


def compute_tags_topicos(validacion: dict | None) -> str:
    if validacion is None:
        return ""
    tags = validacion.get("tags_topicos")
    if not isinstance(tags, list):
        return ""
    return "; ".join(str(t) for t in tags)


def scalar(obj: Any, path: str) -> str:
    v = get_path(obj, path)
    if v is None:
        return ""
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (int, float)):
        return str(v)
    return str(v)


COLUMN_DEFS: list[tuple[str, Callable[[dict, dict | None], str]]] = [
    ("schema_version", lambda n, e: scalar(n, "schema_version")),
    ("codebook_version", lambda n, e: scalar(n, "codebook_version")),
    ("nota_id", lambda n, e: scalar(n, "nota.nota_id")),
    ("nota_archivo_fuente", lambda n, e: scalar(n, "nota.archivo_fuente")),
    ("nota_fuente", lambda n, e: scalar(n, "nota.fuente")),
    ("nota_fecha_publicacion", lambda n, e: scalar(n, "nota.fecha_publicacion")),
    ("nota_titulo", lambda n, e: scalar(n, "nota.titulo")),
    ("nota_subtitulo", lambda n, e: scalar(n, "nota.subtitulo")),
    ("texto_original", lambda n, e: scalar(n, "nota.texto_original")),
    ("meta_modelo", lambda n, e: scalar(n, "metadatos_extraccion.modelo")),
    ("meta_proveedor", lambda n, e: scalar(n, "metadatos_extraccion.proveedor")),
    ("meta_fecha_extraccion", lambda n, e: scalar(n, "metadatos_extraccion.fecha_extraccion")),
    ("meta_prompt_version", lambda n, e: scalar(n, "metadatos_extraccion.prompt_version")),
    ("meta_extractor_tipo", lambda n, e: scalar(n, "metadatos_extraccion.extractor_tipo")),
    ("meta_modo_procesamiento", lambda n, e: scalar(n, "metadatos_extraccion.modo_procesamiento")),
    ("meta_estado_validacion_humana", lambda n, e: scalar(n, "metadatos_extraccion.estado_validacion_humana")),
    ("meta_observaciones_metodologicas", lambda n, e: scalar(n, "metadatos_extraccion.observaciones_metodologicas")),
    ("meta_codebook_version", lambda n, e: scalar(n, "metadatos_extraccion.codebook_version")),
    ("tiene_eventos_protesta", lambda n, e: scalar(n, "extraccion.tiene_eventos_protesta")),
    ("total_eventos_protesta", lambda n, e: scalar(n, "extraccion.total_eventos_protesta")),
    ("extraccion_observaciones_extraccion", lambda n, e: scalar(n, "extraccion.observaciones_extraccion")),
    ("estado_validacion_humana", lambda n, e: compute_estado(n.get("validacion_humana"))),
    ("tags_topicos", lambda n, e: compute_tags_topicos(n.get("validacion_humana"))),
    ("modificado_por", lambda n, e: compute_modificado_por(n.get("validacion_humana"))),
    ("validacion_modificada", lambda n, e: scalar(n, "validacion_humana.modificada")),
    ("validacion_usuario_validador", lambda n, e: scalar(n, "validacion_humana.usuario_validador")),
    ("validacion_fecha_revision", lambda n, e: scalar(n, "validacion_humana.fecha_revision")),
    ("validacion_observaciones", lambda n, e: scalar(n, "validacion_humana.observaciones")),
    ("evento_id", lambda n, e: scalar(e, "evento_id") if e else ""),
    ("evento_numero", lambda n, e: scalar(e, "evento_numero") if e else ""),
    ("es_evento_protesta", lambda n, e: scalar(e, "es_evento_protesta") if e else ""),
    ("observaciones_evento", lambda n, e: scalar(e, "observaciones_evento") if e else ""),
    ("delim_descripcion_sintetica", lambda n, e: scalar(e, "delimitacion_evento.descripcion_sintetica") if e else ""),
    ("delim_criterio_delimitacion", lambda n, e: scalar(e, "delimitacion_evento.criterio_delimitacion") if e else ""),
    ("delim_es_accion_principal_con_complementarias", lambda n, e: scalar(e, "delimitacion_evento.es_accion_principal_con_complementarias") if e else ""),
    ("delim_cita_textual_evento", lambda n, e: scalar(e, "delimitacion_evento.cita_textual_evento") if e else ""),
    ("delim_razonamiento", lambda n, e: scalar(e, "delimitacion_evento.razonamiento") if e else ""),
    ("tipo_temporal", lambda n, e: scalar(e, "temporalidad.tipo_temporal") if e else ""),
    ("tempo_verbal", lambda n, e: scalar(e, "temporalidad.tempo_verbal") if e else ""),
    ("fecha_inicio_valor", lambda n, e: scalar(e, "temporalidad.fecha_inicio.valor") if e else ""),
    ("fecha_inicio_tipo_inferencia", lambda n, e: scalar(e, "temporalidad.fecha_inicio.tipo_inferencia") if e else ""),
    ("fecha_inicio_cita_textual", lambda n, e: scalar(e, "temporalidad.fecha_inicio.cita_textual") if e else ""),
    ("fecha_inicio_razonamiento", lambda n, e: scalar(e, "temporalidad.fecha_inicio.razonamiento") if e else ""),
    ("fecha_fin_valor", lambda n, e: scalar(e, "temporalidad.fecha_fin.valor") if e else ""),
    ("fecha_fin_tipo_inferencia", lambda n, e: scalar(e, "temporalidad.fecha_fin.tipo_inferencia") if e else ""),
    ("fecha_fin_cita_textual", lambda n, e: scalar(e, "temporalidad.fecha_fin.cita_textual") if e else ""),
    ("fecha_fin_razonamiento", lambda n, e: scalar(e, "temporalidad.fecha_fin.razonamiento") if e else ""),
    ("expresion_temporal_textual", lambda n, e: scalar(e, "temporalidad.expresion_temporal_textual") if e else ""),
    ("fecha_publicacion_usada_como_referencia", lambda n, e: scalar(e, "temporalidad.fecha_publicacion_usada_como_referencia") if e else ""),
    ("temporalidad_razonamiento", lambda n, e: scalar(e, "temporalidad.razonamiento") if e else ""),
    ("accion_descripcion_textual", lambda n, e: scalar(e, "accion.descripcion_textual") if e else ""),
    ("accion_formato_principal_cita_textual", lambda n, e: scalar(e, "accion.formato_principal.cita_textual") if e else ""),
    ("accion_valor_textual", lambda n, e: scalar(e, "accion.formato_principal.valor_textual") if e else ""),
    ("accion_categoria", lambda n, e: scalar(e, "accion.formato_principal.categoria") if e else ""),
    ("accion_formato_principal_subtipo_textual", lambda n, e: scalar(e, "accion.formato_principal.subtipo_textual") if e else ""),
    ("accion_formato_principal_razonamiento", lambda n, e: scalar(e, "accion.formato_principal.razonamiento") if e else ""),
    ("accion_formatos_complementarios", lambda n, e: array_field(e.get("accion", {}), "formatos_complementarios") if e else ""),
    ("accion_razonamiento", lambda n, e: scalar(e, "accion.razonamiento") if e else ""),
    ("sujetos", lambda n, e: array_field(e, "sujetos") if e else ""),
    ("motivos", lambda n, e: array_field(e, "motivos") if e else ""),
    ("demandas", lambda n, e: array_field(e, "demandas") if e else ""),
    ("contra_quien", lambda n, e: array_field(e, "contra_quien") if e else ""),
    ("lugares", lambda n, e: array_field(e, "lugares") if e else ""),
    ("alcance_categoria", lambda n, e: scalar(e, "alcance.categoria") if e else ""),
    ("alcance_cita_textual", lambda n, e: scalar(e, "alcance.cita_textual") if e else ""),
    ("alcance_razonamiento", lambda n, e: scalar(e, "alcance.razonamiento") if e else ""),
    ("cantidad_hay_mencionada", lambda n, e: scalar(e, "cantidad_participantes.hay_cantidad_mencionada") if e else ""),
    ("cantidad_valor", lambda n, e: scalar(e, "cantidad_participantes.valor") if e else ""),
    ("cantidad_valor_textual", lambda n, e: scalar(e, "cantidad_participantes.valor_textual") if e else ""),
    ("cantidad_es_aproximada", lambda n, e: scalar(e, "cantidad_participantes.es_aproximada") if e else ""),
    ("cantidad_cita_textual", lambda n, e: scalar(e, "cantidad_participantes.cita_textual") if e else ""),
    ("cantidad_fuente_de_la_cifra", lambda n, e: scalar(e, "cantidad_participantes.fuente_de_la_cifra") if e else ""),
    ("cantidad_razonamiento", lambda n, e: scalar(e, "cantidad_participantes.razonamiento") if e else ""),
    ("incidentes_represion_presencia", lambda n, e: scalar(e, "incidentes.represion.presencia") if e else ""),
    ("incidentes_represion_descripcion", lambda n, e: scalar(e, "incidentes.represion.descripcion") if e else ""),
    ("incidentes_represion_cita_textual", lambda n, e: scalar(e, "incidentes.represion.cita_textual") if e else ""),
    ("incidentes_represion_razonamiento", lambda n, e: scalar(e, "incidentes.represion.razonamiento") if e else ""),
    ("incidentes_enfrentamiento_presencia", lambda n, e: scalar(e, "incidentes.enfrentamiento.presencia") if e else ""),
    ("incidentes_enfrentamiento_descripcion", lambda n, e: scalar(e, "incidentes.enfrentamiento.descripcion") if e else ""),
    ("incidentes_enfrentamiento_cita_textual", lambda n, e: scalar(e, "incidentes.enfrentamiento.cita_textual") if e else ""),
    ("incidentes_enfrentamiento_razonamiento", lambda n, e: scalar(e, "incidentes.enfrentamiento.razonamiento") if e else ""),
    ("incidentes_detenidos_presencia", lambda n, e: scalar(e, "incidentes.detenidos.presencia") if e else ""),
    ("incidentes_detenidos_valor", lambda n, e: scalar(e, "incidentes.detenidos.valor") if e else ""),
    ("incidentes_detenidos_valor_textual", lambda n, e: scalar(e, "incidentes.detenidos.valor_textual") if e else ""),
    ("incidentes_detenidos_cita_textual", lambda n, e: scalar(e, "incidentes.detenidos.cita_textual") if e else ""),
    ("incidentes_detenidos_razonamiento", lambda n, e: scalar(e, "incidentes.detenidos.razonamiento") if e else ""),
    ("incidentes_heridos_presencia", lambda n, e: scalar(e, "incidentes.heridos.presencia") if e else ""),
    ("incidentes_heridos_valor", lambda n, e: scalar(e, "incidentes.heridos.valor") if e else ""),
    ("incidentes_heridos_valor_textual", lambda n, e: scalar(e, "incidentes.heridos.valor_textual") if e else ""),
    ("incidentes_heridos_cita_textual", lambda n, e: scalar(e, "incidentes.heridos.cita_textual") if e else ""),
    ("incidentes_heridos_razonamiento", lambda n, e: scalar(e, "incidentes.heridos.razonamiento") if e else ""),
    ("incidentes_muertos_presencia", lambda n, e: scalar(e, "incidentes.muertos.presencia") if e else ""),
    ("incidentes_muertos_valor", lambda n, e: scalar(e, "incidentes.muertos.valor") if e else ""),
    ("incidentes_muertos_valor_textual", lambda n, e: scalar(e, "incidentes.muertos.valor_textual") if e else ""),
    ("incidentes_muertos_cita_textual", lambda n, e: scalar(e, "incidentes.muertos.cita_textual") if e else ""),
    ("incidentes_muertos_razonamiento", lambda n, e: scalar(e, "incidentes.muertos.razonamiento") if e else ""),
    ("individuos_hay_nombrados", lambda n, e: scalar(e, "individuos_nombrados.hay_individuos_nombrados") if e else ""),
    ("individuos_personas", lambda n, e: array_field(e.get("individuos_nombrados", {}), "personas") if e else ""),
    ("voces_hay_protagonistas", lambda n, e: scalar(e, "voces_protagonistas.hay_voces_protagonistas") if e else ""),
    ("voces_citas", lambda n, e: array_field(e.get("voces_protagonistas", {}), "citas") if e else ""),
    ("control_campos_inferidos", lambda n, e: array_field(e.get("control_extraccion", {}), "campos_inferidos") if e else ""),
    ("control_ambiguedades", lambda n, e: array_field(e.get("control_extraccion", {}), "ambiguedades") if e else ""),
    ("control_advertencias_atomicidad", lambda n, e: array_field(e.get("control_extraccion", {}), "advertencias_atomicidad") if e else ""),
    ("control_requiere_revision_humana", lambda n, e: scalar(e, "control_extraccion.requiere_revision_humana") if e else ""),
    ("control_razones_revision", lambda n, e: array_field(e.get("control_extraccion", {}), "razones_revision") if e else ""),
]


def build_rows(src_path: Path) -> list[list[str]]:
    rows: list[list[str]] = []
    with src_path.open("r", encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            nota_meta = json.loads(line)
            eventos = nota_meta.get("extraccion", {}).get("eventos_protesta") or []
            if not eventos:
                row = [extractor(nota_meta, None) for _, extractor in COLUMN_DEFS]
                rows.append(row)
                continue
            for evento in eventos:
                row = [extractor(nota_meta, evento) for _, extractor in COLUMN_DEFS]
                rows.append(row)
    return rows


def write_workbook(rows: list[list[str]], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "entrenamiento"
    ws.append([name for name, _ in COLUMN_DEFS])
    for row in rows:
        ws.append(row)

    for idx, (name, _) in enumerate(COLUMN_DEFS, start=1):
        if name == "texto_original":
            width = 80
        elif name in {"nota_id", "nota_archivo_fuente"}:
            width = 52
        elif name in {"delim_cita_textual_evento", "delim_descripcion_sintetica", "individuos_personas", "voces_citas"}:
            width = 60
        elif name in {"motivos", "demandas", "sujetos", "lugares", "contra_quien"}:
            width = 70
        elif name in {"accion_formatos_complementarios"}:
            width = 50
        else:
            width = 22
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "B2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
    wb.save(out_path)


def main() -> None:
    rows = build_rows(SRC)
    write_workbook(rows, OUT)
    print(f"escritas {len(rows)} filas en {OUT.name}")
    print(f"columnas: {len(COLUMN_DEFS)}")


if __name__ == "__main__":
    main()